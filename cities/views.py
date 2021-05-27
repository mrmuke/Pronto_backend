from django.shortcuts import render
from nltk.corpus import wordnet
from openpyxl import load_workbook
from rest_framework import generics
from rest_framework.response import Response
import itertools
import operator
import requests
import wikipedia
import numpy as np
import json


#Scrape from lonelyplanet.com and use edit distance
#compare synonyms with edit distance to descriptions
def levenshtein(seq1, seq2):
    size_x = len(seq1) + 1
    size_y = len(seq2) + 1
    matrix = np.zeros ((size_x, size_y))
    for x in range(size_x):
        matrix [x, 0] = x
    for y in range(size_y):
        matrix [0, y] = y

    for x in range(1, size_x):
        for y in range(1, size_y):
            if seq1[x-1] == seq2[y-1]:
                matrix [x,y] = min(
                    matrix[x-1, y] + 1,
                    matrix[x-1, y-1],
                    matrix[x, y-1] + 1
                )
            else:
                matrix [x,y] = min(
                    matrix[x-1,y] + 1,
                    matrix[x-1,y-1] + 1,
                    matrix[x,y-1] + 1
                )
    return (matrix[size_x - 1, size_y - 1])
""" 
def get_word_synonyms_from_sent(word, sent):
    word_synonyms = []
    for synset in wordnet.synsets(word):
        for lemma in synset.lemma_names():
            if lemma in sent:

                word_synonyms.append(lemma)
    return word_synonyms

class AdjectivesDestination(generics.RetrieveAPIView):

    def get(self, *args, **kwargs):
        #adjs = ["warm","beautiful","cultural"]
        adjs=eval(self.request.GET.get("adjs"))
        
        wb = load_workbook("AdjectivesCityTravel.xlsx", data_only=True)
        destinations ={}
        for i in range(2,22):
            adjectives = []
            
            for j in range(2,7):
                val=wb.active.cell(i,j).value
                if(val):
                    adjectives.append(val.lower())
            destinations[wb.active.cell(i,1).value]=adjectives



        scores={}
        for place in destinations:
            for adj in adjs:
                print(adj)
                print(destinations[place])
                word_synonyms=get_word_synonyms_from_sent(adj,destinations[place])
                scores[place]=scores.get(place, 0) + len(word_synonyms)

        return Response(max(scores.items(), key=operator.itemgetter(1))[0])
 """

def get_top_rating(synonyms, target):
    top=0
    for synonym in synonyms:
        for word in target:
            levDis=levenshtein(synonym,word)
            bigger = max(len(synonym), len(word))
            ratio = float(bigger - levDis) / bigger
            if(ratio>top):
                top=ratio
    return top

#fix and test
class DictionaryAdjectivesDestination(generics.RetrieveAPIView):

    def get(self, *args, **kwargs):
        #adjs = ["warm","beautiful","cultural"]
        adjs=eval(self.request.GET.get("adjs"))

        
        destinations ={}
        with open('./cityAdjectives.json') as f:
            destinations = json.load(f)


        scores={}
        synonyms={}
        for adj in adjs:
            synonyms[adj]=[adj]
            r= requests.get("https://www.dictionaryapi.com/api/v3/references/thesaurus/json/{}?key=fdae46e0-553d-45ea-8d24-860b4625495f".format(adj))
                
            syns =r.json()[0]["meta"]["syns"]
            synonyms[adj]+=syns
        for place in destinations:
            for adj in adjs:
                #store synonms in dict
                
                
                word_synonym_count=get_top_rating(synonyms[adj],destinations[place])
                scores[place]=scores.get(place, 0) +word_synonym_count

        return Response(max(scores.items(), key=operator.itemgetter(1))[0])
        



class GetCity(generics.RetrieveAPIView):

    def get(self, *args, **kwargs):
        name=self.request.GET.get("city")
        name=name.replace("-", " ")
        print(name)
        city={}
        with open('./cityInformation.json') as f:
            city = json.load(f)[name]
        
#embed google maps picture
        return Response(city)

